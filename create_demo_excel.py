"""
Create a demo Excel file that mimics the Arla file structure described in requirements.
This demonstrates the different table scenarios.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Media Plan"

# Scenario 1: Table without big header (Media Type to 100%)
# Row 1: Empty (no big header)
# Row 2: Headers
ws['A2'] = 'Media Type'
ws['B2'] = 'Channel'
ws['C2'] = 'Budget'
ws['D2'] = '0%'
ws['E2'] = '25%'
ws['F2'] = '50%'
ws['G2'] = '75%'
ws['H2'] = '100%'

# Data rows for table 1
ws['A3'] = 'Digital'
ws['B3'] = 'Social Media'
ws['C3'] = 50000
ws['D3'] = 0
ws['E3'] = 12500
ws['F3'] = 25000
ws['G3'] = 37500
ws['H3'] = 50000

ws['A4'] = 'TV'
ws['B4'] = 'Prime Time'
ws['C4'] = 100000
ws['D4'] = 0
ws['E4'] = 25000
ws['F4'] = 50000
ws['G4'] = 75000
ws['H4'] = 100000

# Empty column I to separate tables

# Scenario 2: Table with merged big header (Jan-Dec)
# Row 1: Big header "Budget (Jan-Dec)"
ws.merge_cells('J1:U1')
ws['J1'] = 'Budget (Jan-Dec)'
ws['J1'].font = Font(bold=True, size=12)
ws['J1'].alignment = Alignment(horizontal='center')

# Row 2: Month sub-headers
months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
for i, month in enumerate(months):
    col = get_column_letter(10 + i)  # Starting from column J (10)
    ws[f'{col}2'] = month

# Data for table 2
budget_data = [45000, 47000, 52000, 50000, 55000, 58000, 60000, 58000, 56000, 54000, 62000, 65000]
for i, value in enumerate(budget_data):
    col = get_column_letter(10 + i)
    ws[f'{col}3'] = value

actual_data = [44500, 46800, 51500, 49800, 54200, 57500, 59800, 57900, 55600, 53800, 61500, 64200]
for i, value in enumerate(actual_data):
    col = get_column_letter(10 + i)
    ws[f'{col}4'] = value

# Create second sheet for more complex scenarios
ws2 = wb.create_sheet("Campaign Details")

# Scenario 3: Big header "Buy Specifics"
ws2.merge_cells('A1:E1')
ws2['A1'] = 'Buy Specifics'
ws2['A1'].font = Font(bold=True, size=12)
ws2['A1'].alignment = Alignment(horizontal='center')

# Sub-headers under Buy Specifics
ws2['A2'] = 'Target Audience'
ws2['B2'] = 'Age Range'
ws2['C2'] = 'Gender'
ws2['D2'] = 'Location'
ws2['E2'] = 'If excluded reason'

# Data
ws2['A3'] = 'Adults'
ws2['B3'] = '25-45'
ws2['C3'] = 'All'
ws2['D3'] = 'Urban'
ws2['E3'] = 'N/A'

# Empty column F

# Scenario 4: Multiple big headers without merging (Seasonality)
# Row 1: Big headers (not merged, but detected by pattern)
ws2['G1'] = 'Cost Guarantee'
ws2['H1'] = 'Budget Variation Index'
ws2['M1'] = 'Seasonality'

# Row 2: Sub-headers
ws2['G2'] = 'L3 Benchmark'
ws2['H2'] = 'Jan'
ws2['I2'] = 'Feb'
ws2['J2'] = 'Mar'
ws2['K2'] = 'Apr'
ws2['L2'] = 'May'
ws2['M2'] = 'Jan'
ws2['N2'] = 'Feb'
ws2['O2'] = 'Mar'
ws2['P2'] = 'Apr'
ws2['Q2'] = 'May'

# Data
ws2['G3'] = 15000
ws2['H3'] = 1.05
ws2['I3'] = 1.02
ws2['J3'] = 1.08
ws2['K3'] = 1.03
ws2['L3'] = 1.10
ws2['M3'] = 0.95
ws2['N3'] = 1.00
ws2['O3'] = 1.15
ws2['P3'] = 1.05
ws2['Q3'] = 1.20

# Save the demo file
output_path = 'data/Demo_Arla_Example.xlsx'
wb.save(output_path)
print(f"✅ Demo Excel file created: {output_path}")
print("\nThis file demonstrates:")
print("  1. Table without big header (Media Type to 100%)")
print("  2. Table with merged big header (Budget Jan-Dec)")
print("  3. Buy Specifics table with merged header")
print("  4. Multiple tables with inferred big headers (Seasonality)")
