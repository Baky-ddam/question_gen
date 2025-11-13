"""
Header detection logic for Excel tables.
Handles merged headers, big headers, and sub-headers.
"""

import re
from typing import List, Tuple, Optional, Dict, Set
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter


class HeaderDetector:
    """Detects and classifies headers in Excel sheets."""

    # Common patterns that indicate a sequence header (big header exists)
    MONTH_PATTERNS = [
        r'\b(jan|january)\b.*\b(dec|december)\b',
        r'\bjan\b.*\bdec\b',
        r'\b(january|february|march|april|may|june|july|august|september|october|november|december)\b'
    ]

    PERCENTAGE_PATTERNS = [
        r'\b0%\b.*\b100%\b',
        r'\b\d+%\b.*\b\d+%\b'
    ]

    NUMERIC_SEQUENCE_PATTERNS = [
        r'\b\d+\b.*\b\d+\b'  # Any numeric sequence
    ]

    def __init__(self):
        self.merged_ranges: Dict[str, List] = {}

    def detect_merged_cells(self, worksheet: Worksheet, max_rows: int = 10) -> Dict[Tuple[int, int], Tuple[int, int, int, int]]:
        """
        Detect merged cells in the worksheet (typically in header rows).
        Returns dict mapping (row, col) -> (start_row, start_col, end_row, end_col)
        """
        merged_cells_map = {}

        for merged_range in worksheet.merged_cells.ranges:
            min_row, min_col = merged_range.min_row, merged_range.min_col
            max_row, max_col = merged_range.max_row, merged_range.max_col

            # Only consider merged cells in header area
            if min_row <= max_rows:
                # Map all cells in the merged range to the same info
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        merged_cells_map[(row, col)] = (min_row, min_col, max_row, max_col)

        return merged_cells_map

    def is_month_sequence(self, headers: List[str]) -> bool:
        """Check if headers represent a month sequence (Jan-Dec)."""
        headers_str = ' '.join(str(h).lower() for h in headers if h)

        for pattern in self.MONTH_PATTERNS:
            if re.search(pattern, headers_str, re.IGNORECASE):
                return True

        # Check if we have consecutive months
        months = ['jan', 'feb', 'mar', 'apr', 'may', 'jun',
                  'jul', 'aug', 'sep', 'oct', 'nov', 'dec']
        month_count = sum(1 for h in headers if any(m in str(h).lower() for m in months))

        return month_count >= 3  # At least 3 consecutive months

    def is_percentage_sequence(self, headers: List[str]) -> bool:
        """Check if headers represent a percentage sequence."""
        headers_str = ' '.join(str(h) for h in headers if h)

        for pattern in self.PERCENTAGE_PATTERNS:
            if re.search(pattern, headers_str):
                return True

        # Check if most headers are percentages
        pct_count = sum(1 for h in headers if h and '%' in str(h))
        return pct_count >= 2 and pct_count >= len(headers) * 0.5

    def is_numeric_sequence(self, headers: List[str]) -> bool:
        """Check if headers are numeric sequence."""
        try:
            numbers = []
            for h in headers:
                if h and str(h).strip():
                    # Try to extract number
                    match = re.search(r'\d+', str(h))
                    if match:
                        numbers.append(int(match.group()))

            # Check if we have at least 3 numbers
            return len(numbers) >= 3
        except:
            return False

    def detect_sequence_pattern(self, headers: List[str]) -> bool:
        """
        Detect if headers follow a pattern that suggests a big header exists.
        Examples: Jan-Dec, 0%-100%, numeric sequences
        """
        if not headers:
            return False

        return (self.is_month_sequence(headers) or
                self.is_percentage_sequence(headers) or
                self.is_numeric_sequence(headers))

    def find_big_header_for_sequence(self, worksheet: Worksheet, row_idx: int,
                                     start_col: int, end_col: int) -> Optional[str]:
        """
        Find the big header for a sequence by looking at the row above.
        """
        if row_idx <= 1:
            return None

        # Check the row above for merged cells or a single header
        above_row = row_idx - 1

        # Get value from the cell above the start column
        cell = worksheet.cell(row=above_row, column=start_col)

        if cell.value and str(cell.value).strip():
            return str(cell.value).strip()

        return None

    def is_unnamed_column(self, header: str) -> bool:
        """Check if a column header is unnamed or empty."""
        if not header:
            return True

        header_str = str(header).strip().lower()

        # Common patterns for unnamed columns
        unnamed_patterns = [
            'unnamed',
            'column',
            r'^col\d+$',
            r'^\d+$',  # Just a number
            ''
        ]

        for pattern in unnamed_patterns:
            if re.match(pattern, header_str):
                return True

        return False

    def detect_table_boundaries(self, headers: List[str]) -> List[Tuple[int, int]]:
        """
        Detect table boundaries based on Unnamed columns.
        Returns list of (start_idx, end_idx) tuples for each table.
        """
        if not headers:
            return []

        tables = []
        start_idx = None

        for i, header in enumerate(headers):
            is_unnamed = self.is_unnamed_column(header)

            if not is_unnamed and start_idx is None:
                # Start of a new table
                start_idx = i
            elif is_unnamed and start_idx is not None:
                # End of current table
                tables.append((start_idx, i - 1))
                start_idx = None

        # Add the last table if it extends to the end
        if start_idx is not None:
            tables.append((start_idx, len(headers) - 1))

        return tables

    def extract_headers_with_hierarchy(self, worksheet: Worksheet,
                                       header_row: int = 1,
                                       big_header_row: Optional[int] = None) -> List[Tuple[str, Optional[str], int]]:
        """
        Extract headers with hierarchy (big_header, sub_header, column_index).
        Returns list of tuples: (sub_header, big_header, column_index)
        """
        if big_header_row is None:
            big_header_row = header_row - 1 if header_row > 1 else None

        headers_with_hierarchy = []
        merged_cells = self.detect_merged_cells(worksheet, max_rows=header_row)

        max_col = worksheet.max_column

        for col_idx in range(1, max_col + 1):
            # Get sub-header (main header row)
            sub_header_cell = worksheet.cell(row=header_row, column=col_idx)
            sub_header = str(sub_header_cell.value).strip() if sub_header_cell.value else ""

            # Try to get big header
            big_header = None

            if big_header_row:
                # Check if there's a merged cell above
                merge_key = (big_header_row, col_idx)
                if merge_key in merged_cells:
                    merge_info = merged_cells[merge_key]
                    min_row, min_col, max_row, max_col_merge = merge_info
                    # Get value from the merged cell
                    big_header_cell = worksheet.cell(row=min_row, column=min_col)
                    if big_header_cell.value:
                        big_header = str(big_header_cell.value).strip()
                else:
                    # Check for non-merged big header
                    big_header_cell = worksheet.cell(row=big_header_row, column=col_idx)
                    if big_header_cell.value:
                        big_header = str(big_header_cell.value).strip()

            headers_with_hierarchy.append((sub_header, big_header, col_idx))

        return headers_with_hierarchy

    def group_columns_by_big_header(self, headers_with_hierarchy: List[Tuple[str, Optional[str], int]]) -> Dict[Optional[str], List[Tuple[str, int]]]:
        """
        Group columns by their big header.
        Returns dict: big_header -> [(sub_header, col_idx), ...]
        """
        groups = {}

        for sub_header, big_header, col_idx in headers_with_hierarchy:
            if big_header not in groups:
                groups[big_header] = []
            groups[big_header].append((sub_header, col_idx))

        return groups

    def infer_big_headers_from_patterns(self, worksheet: Worksheet,
                                        header_row: int,
                                        headers: List[str]) -> Dict[int, str]:
        """
        Infer big headers from patterns even if they're not merged.
        Returns dict: column_index -> big_header_name
        """
        inferred_big_headers = {}

        if not headers:
            return inferred_big_headers

        # Group consecutive columns that follow a pattern
        i = 0
        while i < len(headers):
            # Look ahead to find a sequence
            sequence_start = i
            sequence_headers = []

            # Collect potential sequence
            for j in range(i, min(i + 20, len(headers))):  # Look ahead up to 20 columns
                if not self.is_unnamed_column(headers[j]):
                    sequence_headers.append(headers[j])
                else:
                    break

            # Check if this is a pattern sequence
            if self.detect_sequence_pattern(sequence_headers):
                # Try to find big header above
                big_header = self.find_big_header_for_sequence(
                    worksheet, header_row, sequence_start + 1, sequence_start + len(sequence_headers)
                )

                if not big_header:
                    # Infer big header name from the pattern itself
                    if self.is_month_sequence(sequence_headers):
                        big_header = "Monthly Data"
                    elif self.is_percentage_sequence(sequence_headers):
                        big_header = "Percentage Distribution"
                    else:
                        big_header = "Numeric Sequence"

                # Assign this big header to all columns in the sequence
                for k in range(len(sequence_headers)):
                    inferred_big_headers[sequence_start + k] = big_header

                i += len(sequence_headers)
            else:
                i += 1

        return inferred_big_headers
