"""
Main Excel table extraction engine.
Extracts tables from Excel files with support for merged headers and multi-level headers.
"""

import os
from typing import List, Dict, Optional, Any
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

from .table_structure import (
    TableStructure, ColumnMapping, SheetTables, WorkbookTables,
    HeaderType
)
from .header_detector import HeaderDetector


class ExcelTableExtractor:
    """
    Main class for extracting tables from Excel files.

    Features:
    - Detects merged headers (big headers)
    - Tags sub-headers with their big header
    - Detects table boundaries using Unnamed columns
    - Handles patterns like Jan-Dec, percentages, etc.
    """

    def __init__(self, header_row: int = 2, big_header_row: Optional[int] = 1):
        """
        Initialize the table extractor.

        Args:
            header_row: Row number for sub-headers (1-indexed)
            big_header_row: Row number for big headers (1-indexed), None if no big headers
        """
        self.header_row = header_row
        self.big_header_row = big_header_row
        self.header_detector = HeaderDetector()

    def extract_from_file(self, filepath: str, sheet_names: Optional[List[str]] = None) -> WorkbookTables:
        """
        Extract all tables from an Excel file.

        Args:
            filepath: Path to Excel file
            sheet_names: List of sheet names to process (None = all sheets)

        Returns:
            WorkbookTables containing all extracted tables
        """
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"File not found: {filepath}")

        workbook = load_workbook(filepath, data_only=True)
        filename = os.path.basename(filepath)
        workbook_tables = WorkbookTables(filename=filename)

        # Process sheets
        sheets_to_process = sheet_names if sheet_names else workbook.sheetnames

        for sheet_name in sheets_to_process:
            if sheet_name not in workbook.sheetnames:
                print(f"Warning: Sheet '{sheet_name}' not found in workbook")
                continue

            worksheet = workbook[sheet_name]
            sheet_tables = self.extract_from_sheet(worksheet, sheet_name)
            workbook_tables.add_sheet(sheet_tables)

        workbook.close()
        return workbook_tables

    def extract_from_sheet(self, worksheet: Worksheet, sheet_name: str) -> SheetTables:
        """
        Extract all tables from a single sheet.

        Args:
            worksheet: Openpyxl worksheet object
            sheet_name: Name of the sheet

        Returns:
            SheetTables containing all tables from this sheet
        """
        sheet_tables = SheetTables(sheet_name=sheet_name)

        # Extract headers with hierarchy
        headers_with_hierarchy = self.header_detector.extract_headers_with_hierarchy(
            worksheet,
            header_row=self.header_row,
            big_header_row=self.big_header_row
        )

        if not headers_with_hierarchy:
            return sheet_tables

        # Extract just sub-headers for boundary detection
        sub_headers = [h[0] for h in headers_with_hierarchy]

        # Detect table boundaries
        table_boundaries = self.header_detector.detect_table_boundaries(sub_headers)

        if not table_boundaries:
            # No clear boundaries, treat entire sheet as one table
            table_boundaries = [(0, len(sub_headers) - 1)]

        # Infer big headers from patterns if not explicitly merged
        inferred_big_headers = self.header_detector.infer_big_headers_from_patterns(
            worksheet, self.header_row, sub_headers
        )

        # Extract each table
        for table_id, (start_col_idx, end_col_idx) in enumerate(table_boundaries, start=1):
            table = self._extract_single_table(
                worksheet=worksheet,
                sheet_name=sheet_name,
                table_id=table_id,
                headers_with_hierarchy=headers_with_hierarchy,
                start_col_idx=start_col_idx,
                end_col_idx=end_col_idx,
                inferred_big_headers=inferred_big_headers
            )
            sheet_tables.add_table(table)

        return sheet_tables

    def _extract_single_table(self,
                              worksheet: Worksheet,
                              sheet_name: str,
                              table_id: int,
                              headers_with_hierarchy: List[tuple],
                              start_col_idx: int,
                              end_col_idx: int,
                              inferred_big_headers: Dict[int, str]) -> TableStructure:
        """
        Extract a single table from the sheet.

        Args:
            worksheet: Openpyxl worksheet
            sheet_name: Sheet name
            table_id: Table identifier
            headers_with_hierarchy: List of (sub_header, big_header, col_idx)
            start_col_idx: Starting column index (0-based)
            end_col_idx: Ending column index (0-based)
            inferred_big_headers: Dict of inferred big headers

        Returns:
            TableStructure with extracted data
        """
        # Extract columns for this table
        table_headers = headers_with_hierarchy[start_col_idx:end_col_idx + 1]

        # Create column mappings
        columns = []
        has_merged_headers = False

        for sub_header, big_header, col_idx in table_headers:
            # Use inferred big header if available and no explicit big header
            if not big_header and (start_col_idx + len(columns)) in inferred_big_headers:
                big_header = inferred_big_headers[start_col_idx + len(columns)]

            if big_header:
                has_merged_headers = True
                header_type = HeaderType.SUB_HEADER
            else:
                header_type = HeaderType.STANDALONE

            column_mapping = ColumnMapping(
                sheet_name=sheet_name,
                big_header=big_header,
                sub_header=sub_header,
                column_index=col_idx,
                column_letter=get_column_letter(col_idx),
                header_type=header_type
            )
            columns.append(column_mapping)

        # Extract data rows
        data_rows = []
        start_data_row = self.header_row + 1
        max_row = worksheet.max_row

        for row_idx in range(start_data_row, max_row + 1):
            row_data = {}
            has_data = False

            for column_mapping in columns:
                cell = worksheet.cell(row=row_idx, column=column_mapping.column_index)
                value = cell.value

                # Use full path as key
                key = column_mapping.get_full_path()
                row_data[key] = value

                if value is not None and str(value).strip():
                    has_data = True

            # Only add row if it has at least some data
            if has_data:
                data_rows.append(row_data)
            elif len(data_rows) > 0:
                # Stop if we hit empty rows after we started collecting data
                break

        # Create table structure
        table = TableStructure(
            sheet_name=sheet_name,
            table_id=table_id,
            start_row=self.header_row,
            end_row=start_data_row + len(data_rows) - 1,
            start_col=columns[0].column_index if columns else 1,
            end_col=columns[-1].column_index if columns else 1,
            columns=columns,
            data=data_rows,
            has_merged_headers=has_merged_headers
        )

        return table

    def extract_summary(self, workbook_tables: WorkbookTables) -> Dict[str, Any]:
        """
        Generate a summary of extracted tables.

        Args:
            workbook_tables: Extracted workbook tables

        Returns:
            Dictionary with summary information
        """
        summary = {
            'filename': workbook_tables.filename,
            'total_sheets': workbook_tables.get_sheet_count(),
            'total_tables': workbook_tables.get_total_tables(),
            'sheets': []
        }

        for sheet_name, sheet_tables in workbook_tables.sheets.items():
            sheet_summary = {
                'name': sheet_name,
                'table_count': sheet_tables.get_table_count(),
                'tables': []
            }

            for table in sheet_tables.tables:
                table_summary = {
                    'table_id': table.table_id,
                    'dimensions': f"{table.get_row_count()} rows x {table.get_column_count()} columns",
                    'has_merged_headers': table.has_merged_headers,
                    'columns': [col.get_full_path() for col in table.columns]
                }
                sheet_summary['tables'].append(table_summary)

            summary['sheets'].append(sheet_summary)

        return summary

    def export_to_json(self, workbook_tables: WorkbookTables, output_path: str):
        """
        Export extracted tables to JSON file.

        Args:
            workbook_tables: Extracted tables
            output_path: Path for output JSON file
        """
        import json

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(workbook_tables.to_dict(), f, indent=2, ensure_ascii=False)

        print(f"Exported to: {output_path}")

    def export_summary(self, workbook_tables: WorkbookTables, output_path: str):
        """
        Export summary to JSON file.

        Args:
            workbook_tables: Extracted tables
            output_path: Path for output JSON file
        """
        import json

        summary = self.extract_summary(workbook_tables)

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(summary, f, indent=2, ensure_ascii=False)

        print(f"Summary exported to: {output_path}")
